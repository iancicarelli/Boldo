from utils.excel_processor import ExcelProcessor
import json
from openpyxl import workbook
import os 
import sys

class CostProcessor:
    def __init__(self, maestra_path, orion_path, bd_ila_path):
        self.maestra_processor = ExcelProcessor(maestra_path)
        self.orion_processor = ExcelProcessor(orion_path)
        self.ila_processor = ExcelProcessor(bd_ila_path) 

    def get_matching_groups(self):
    # Leer los archivos Excel
        self.maestra_processor.read_excel()
        self.orion_processor.read_excel()
    # Obtener las hojas de trabajo
        maestra_sheet = self.maestra_processor.get_sheet()
        orion_sheet = self.orion_processor.get_sheet()
    # Extraer códigos desde la columna A de maestra.xlsx
        maestra_codes = [row[0].value for row in maestra_sheet.iter_rows(min_row=2, max_row=maestra_sheet.max_row, min_col=1, max_col=1) if row[0].value]
    # Diccionario para almacenar coincidencias {codigo: descripcion}
        print(f"maestra_codes = {maestra_codes}")
        matching_dict = {}
    # Buscar coincidencias en BD ORION.xlsx (columna D) y obtener la descripción de la columna F
        for row in orion_sheet.iter_rows(min_row=2, max_row=orion_sheet.max_row, min_col=4, max_col=6):
            orion_code = row[0].value  # Código en columna D
            
            group = row[2].value  # Descripción en columna F
            if orion_code in maestra_codes:
                matching_dict[orion_code] = group  # Guardar la coincidencia
        matching_groups = [matching_dict.get(code, "N/A") for code in maestra_codes]
    
        return matching_groups
  
    def get_ila_values(self):
        self.ila_processor.read_excel()
        ila_sheet = self.ila_processor.get_sheet()

        ila_mapping = {}
        for row in ila_sheet.iter_rows(min_row=2, max_row=ila_sheet.max_row, min_col=2, max_col=5):
            name = row[0].value  # Valor en columna B
            value = row[3].value  # Valor en columna E
            if name and value:  # Solo agregar si ambos valores existen
                ila_mapping[name] = value     
        return ila_mapping

    def match_ila_values_with_groups(self):
        matching_groups = self.get_matching_groups()
        ila_values = self.get_ila_values()
        result = []
        for group in matching_groups:
            if group != "N/A" and group in ila_values:
                result.append(ila_values[group])  # Retorna el valor de ILA si hay coincidencia
            else:
                result.append("N/A")  # Si no hay coincidencia o el valor es "N/A"

        return result
    def get_json_path(self):
        if getattr(sys, 'frozen', False):
        # Si el script está siendo ejecutado desde el .exe
            current_dir = sys._MEIPASS  # Directorio temporal donde PyInstaller coloca los archivos
        else:
            # Si está siendo ejecutado como un script en desarrollo
            current_dir = os.path.dirname(os.path.abspath(__file__))

        # Subir dos niveles para salir de la carpeta "utils" y llegar al directorio raíz del proyecto
        project_root = os.path.dirname(os.path.dirname(current_dir))

        # Crear la ruta correcta al archivo JSON desde el directorio raíz del proyecto o el directorio temporal
        json_path = os.path.join(project_root, "src", "data", "ilas.json")

        # Si estás empaquetando el ejecutable con --add-data, los archivos estarán en el directorio temporal en lugar de "src/data"
        if getattr(sys, 'frozen', False):  # Si se ejecuta como .exe
            json_path = os.path.join(sys._MEIPASS, "data", "ilas.json")

        return json_path

    def compare_with_json(self):
        # Obtener la ruta al archivo JSON usando el método get_json_path
        json_path = self.get_json_path()

        # Leer el archivo JSON
        with open(json_path, "r", encoding="utf-8") as f:
            ila_json = json.load(f)

        # Crear un diccionario de nombres y montos desde el JSON
        json_mapping = {entry["Nombre"]: entry["MONTO"] for entry in ila_json}

        # Obtener la lista de valores de ILA desde match_ila_values_with_groups
        ila_values = self.match_ila_values_with_groups()

        # Comparar y crear la lista de montos correspondientes
        result_montos = []
        for value in ila_values:
            if value != "N/A" and value in json_mapping:
                result_montos.append(json_mapping[value])  # Retorna el MONTO si existe
            else:
                result_montos.append(1.19)  # Retorna el valor por defecto 1.19 si no hay coincidencia

        return result_montos

    def get_column_e_values(self):
        """Extrae los valores de la columna E de maestra.xlsx y los almacena en una lista.
        Si un valor está vacío o contiene '#N/A' o 'N/A', se reemplaza por 'N/A'.
        """
        self.maestra_processor.read_excel()
        maestra_sheet = self.maestra_processor.get_sheet()

        column_e_values = []
        for row in maestra_sheet.iter_rows(min_row=2, max_row=maestra_sheet.max_row, min_col=5, max_col=5):
            value = row[0].value

            if value is None or str(value).strip().upper() in ["#N/A", "N/A"]:
                column_e_values.append("N/A")
            else:
                column_e_values.append(value)

        return column_e_values


    def multiply_lists(self):
        """Multiplica los valores de get_column_e_values() y compare_with_json().
        Si alguno de los valores es 'N/A' o '#N/A', retorna 'N/A' en esa posición.
        """
        column_e_values = self.get_column_e_values()
        json_values = self.compare_with_json()

        # Asegurar que ambas listas tienen la misma longitud
        length = min(len(column_e_values), len(json_values))

        result = []
        for i in range(length):
            val1 = column_e_values[i]
            val2 = json_values[i]

            # Si alguno de los valores no es un número, retornar 'N/A'
            if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
                result.append(round(val1 * val2,2))
            else:
                result.append("N/A")

        return result
    
    def replace_non_numeric_values(self, values):
        """Reemplaza valores no numéricos con 0.0 y asegura que todos los valores sean numéricos."""
        cleaned_values = []
        for v in values:
            if isinstance(v, (int, float)):
                cleaned_values.append(float(v))  # Asegurar que sea número
            else:
                cleaned_values.append(0.0)  # Reemplaza valores no numéricos con 0.0
        return cleaned_values

    def write_to_excel(self, output_path):
        result = self.multiply_lists()  # Obtener los resultados de la multiplicación
        result = self.replace_non_numeric_values(result)  # Asegurar que todos los valores sean numéricos

        wb = workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Escribir los valores en la columna A
        for i, value in enumerate(result, start=1):
            ws.cell(row=i, column=1, value=value)

        try:
            wb.save(output_path)
            print(f"Archivo Excel guardado correctamente en: {output_path}")
        except Exception as e:
            print(f"Error al guardar el archivo Excel: {e}")