from openpyxl import load_workbook, Workbook
from models.out_format import OutFormat
from openpyxl.styles import Font
import shutil
import os
from pathlib import Path
from openpyxl import load_workbook, Workbook

class ExcelProcessor: 
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None

    def copy_excel_file(self,input_path, output_path):
        try:
            # Verificar si la carpeta de salida existe, si no, crearla
            if not os.path.exists(os.path.dirname(output_path)):
                os.makedirs(os.path.dirname(output_path))

            # Copiar el archivo completo (incluyendo hojas, celdas y estilos)
            shutil.copy(input_path, output_path)
            print(f"Archivo copiado correctamente a: {output_path}")
    
        except Exception as e:
         print(f"Error al copiar el archivo: {e}")    

    
    def read_excel(self):
        try:
            # Cargar el libro de trabajo
            self.workbook = load_workbook(self.file_path)
            # Seleccionar la hoja activa
            self.sheet = self.workbook.active
        except Exception as e:
            print(f"Error al cargar el archivo Excel: {e}")

    def write_excel(self, out_formats, output_path):
        try:
            # Crear un nuevo libro de trabajo y una hoja
          # Cargar el archivo existente (que ya tiene las dos hojas)
            workbook = load_workbook(output_path)

            # Crear la tercera hoja
            sheet3 = workbook.create_sheet("FORMATO_ENTRADA_DETALLE")

            # Agregar las cabeceras en la tercera hoja
            self._add_headers(sheet3)
            self._write_additional_data(sheet3)

            # Agregar los datos procesados desde la fila 12
            for row_num, format_obj in enumerate(out_formats, start=12):
                sheet3.cell(row=row_num, column=2, value=format_obj.lin)
                sheet3.cell(row=row_num, column=3, value=format_obj.code)  
                sheet3.cell(row=row_num, column=4, value=format_obj.name)  
                sheet3.cell(row=row_num, column=5, value=format_obj.uxe)
                sheet3.cell(row=row_num, column=6, value=format_obj.cost_neto)

            # Guardar el archivo
            workbook.save(output_path)
            print(f"Archivo generado correctamente en: {output_path}")

        except Exception as e:
            print(f"Error al escribir el archivo Excel: {e}")

    def _add_headers(self, sheet):
        headers = ["Lin", "CODIGO_BARRA", "NOMBRE_ITEM", "UXE", "COSTO_NETO UNITARIO",
                   "MONTO_IVA", "MONTO_IMP_ADIC", "MONTO_DSCTO_UNITARIO",
                   "MONTO_RECARGO_UNITARIO", "MONTO_ITEM"]
        for col_num, header in enumerate(headers, start=2):
            cell = sheet.cell(row=11, column=col_num, value=header)
            cell.font = Font(bold=True)  # Cabecera en negrita
    def _write_additional_data(self, sheet):
            additional_data = {
                "B2": "DESCUENTOS Y/O RECARGOS GLOBAL",
                "B3": "LINEA",
                "C3": "TPO_MOVI",
                "D3": "GLOSA",
                "E3": "TIPO_VALOR",
                "F3": "VALOR",
                "G3": "APLICA_A_EXENTOS",
                "B10": "DETALLE"
                }
    # Write the additional data to the specified cells
            for cell, value in additional_data.items():
                    sheet[cell] = value

    def get_sheet(self):
        return self.sheet
